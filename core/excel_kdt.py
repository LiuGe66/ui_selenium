# -*- coding: utf-8 -*-            
# Author:liu_ge
# @FileName: excel_data.py
# @Time : 2022/11/27 11:15
import logging
from pathlib import Path

import allure
import pytest
from openpyxl.reader.excel import load_workbook
from selenium import webdriver

from core.kdt import KeyWord

logger = logging.getLogger(__name__)


def filter_empty(old_l):
    """过滤空值"""
    new_l = []
    for i in old_l:
        if i:
            new_l.append(i)
    return new_l


def data_by_excel(file):
    """
    从excel中加载用例数据
    :param file:
    :return:
    """
    wb = load_workbook(file)
    logger.info(f"文件{file=},包含了{len(wb.worksheets)}个sheet页")
    suite_dict = {}  # 以套件名称为key,以用例为value
    for ws in wb.worksheets:
        case_dict = {}  # 以名称为key,以步骤为value的字典
        case_name = ""
        for line in ws.iter_rows(values_only=True):
            _id = line[0]
            logger.debug(f"正在处理下一行：{line}")
            if isinstance(_id, int):  # 步骤
                if _id == -1:
                    case_name = line[3]
                    case_dict[case_name] = []  # 以用例名称为Key，创建新的空用例
                elif _id > 0:  # 用例名称
                    case_dict[case_name].append(filter_empty(line))
        logger.info(f"Sheet'{ws.title}',包含了{len(case_dict)}条用例")
        suite_dict[ws.title] = case_dict
    logger.debug(f"生成测试用例{suite_dict=}")
    return suite_dict


def create_case(test_suite: dict, file):
    """
    接收从excel而来的多个测试套件的信息，并生成真正的测试用例
    :return:
    """
    file_path = Path(file)
    filename = file_path.name

    for suite_name, case_dict in test_suite.items():
        @allure.suite(filename)
        class Test:
            @pytest.fixture(autouse=True)
            def init_pytest(self, request):
                self.request = request
                #  把pytest的夹具保存到测试类当中

            @pytest.mark.parametrize('case', case_dict.items(), ids=case_dict.keys())
            def test_(self, case):
                case_name = (case[0])
                logger.warning(f"----------------{suite_name}.{case_name}测试开始----------------")
                print(f"----------------{suite_name}.{case_name}测试开始----------------")
                step_list = case[1]
                kw = KeyWord(request=self.request)  # 不传递driver，传递pytest
                try:
                    for step in step_list:
                        key = step[2]  # 关键字
                        args = step[3:]  # 关键字参数
                        logger.info(f"执行关键字：{key=},{args=}")
                        f = kw.get_kw_method(key)  # 调用关键字
                        try:
                            with allure.step(step[1]):
                                f(*args)
                                allure.attach(
                                    kw.driver.get_screenshot_as_png(),
                                    step[1],
                                    allure.attachment_type.PNG,
                                )
                        except Exception as e:
                            logger.error('关键字调用出错', exc_info=True)
                            # 执行关键字之后截图
                            raise e
                        finally:
                            allure.attach(
                                kw.driver.get_screenshot_as_png(),
                                step[1],
                                allure.attachment_type.PNG,
                            )

                        logger.debug(f"执行关键字：{key=}执行成功")
                    logger.warning(f"----------------{suite_name}.{case_name}测试结束----------------")
                    print(f"----------------{suite_name}.{case_name}测试结束----------------")
                except Exception as e:
                    logger.error(f'{suite_name}.{case_name}测试失败', exc_info=True)
                    print(f'{suite_name}.{case_name}测试失败')
                    raise e

        logger.info(f"生成了测试用例{suite_name}")
        yield Test, suite_name  # 返回了用例和套件名
